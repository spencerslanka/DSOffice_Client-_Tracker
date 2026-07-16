from fastapi import FastAPI, Request, Form
from fastapi.responses import RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import io
import json
import os
import qrcode
from datetime import datetime

from . import database as db

app = FastAPI(title="DS Office Client Tracker")

BASE_DIR = os.path.dirname(__file__)
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
templates.env.filters["tojson"] = json.dumps
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")

# ---------------------------------------------------------------------------
# Admin login
#
# Set ADMIN_USERNAME / ADMIN_PASSWORD / SESSION_SECRET as environment
# variables in production (e.g. on Render). The defaults below are only
# for local testing and are NOT secure — change them before deploying.
# ---------------------------------------------------------------------------
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "cnrnext")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "plds")
SESSION_SECRET = os.environ.get("SESSION_SECRET", "dev-only-secret-change-me")

app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)


def is_logged_in(request: Request) -> bool:
    return request.session.get("is_admin") is True


@app.middleware("http")
async def add_cache_control_header(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


db.init_db()


@app.get("/")
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "active": "home"})


# ---------- Front desk entry ----------
@app.get("/entry")
def entry_form(request: Request):
    return templates.TemplateResponse(
        "entry.html",
        {"request": request, "sections": db.SECTIONS, "token": None, "active": "entry"},
    )


@app.post("/entry")
def entry_submit(
    request: Request,
    name: str = Form(...),
    phone: str = Form(""),
    nic: str = Form(""),
    gender: str = Form(""),
    address: str = Form(""),
    purpose: str = Form(...),
    remarks: str = Form(""),
    sections: list[str] = Form([]),
):
    sections = [s for s in db.SECTIONS if s in sections]
    if not sections:
        sections = db.SECTIONS[:1]
    _id, token = db.create_client(
        name, phone, purpose, sections,
        nic=nic, gender=gender, address=address, remarks=remarks,
    )
    return templates.TemplateResponse(
        "entry.html",
        {
            "request": request, "sections": db.SECTIONS, "token": token,
            "client_name": name, "active": "entry",
        },
    )


# ---------- Self-registration via QR code ----------
@app.get("/register")
def register_form(request: Request):
    return templates.TemplateResponse(
        "register.html",
        {"request": request, "sections": db.SECTIONS, "token": None, "active": "register"},
    )


@app.post("/register")
def register_submit(
    request: Request,
    name: str = Form(...),
    phone: str = Form(""),
    nic: str = Form(""),
    purpose: str = Form(...),
    sections: list[str] = Form([]),
):
    sections = [s for s in db.SECTIONS if s in sections]
    if not sections:
        sections = db.SECTIONS[:1]
    _id, token = db.create_client(name, phone, purpose, sections, nic=nic)
    return templates.TemplateResponse(
        "register.html",
        {
            "request": request, "sections": db.SECTIONS, "token": token,
            "client_name": name, "active": "register",
        },
    )


@app.get("/qr")
def qr_page(request: Request):
    register_url = str(request.base_url).rstrip("/") + "/register"
    return templates.TemplateResponse(
        "qr.html", {"request": request, "active": "qr", "register_url": register_url},
    )


@app.get("/qr-image")
def qr_image(request: Request):
    register_url = str(request.base_url).rstrip("/") + "/register"
    img = qrcode.make(register_url, box_size=10, border=2)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")


@app.get("/qr-feedback")
def qr_feedback_page(request: Request):
    feedback_url = str(request.base_url).rstrip("/") + "/feedback?client=1"
    return templates.TemplateResponse(
        "qr_feedback.html", {"request": request, "active": "qr_feedback", "feedback_url": feedback_url},
    )


@app.get("/qr-feedback-image")
def qr_feedback_image(request: Request):
    feedback_url = str(request.base_url).rstrip("/") + "/feedback?client=1"
    img = qrcode.make(feedback_url, box_size=10, border=2)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")


# ---------- Section officer ----------
@app.get("/section")
def section_view(request: Request, name: str = ""):
    return templates.TemplateResponse(
        "section.html",
        {"request": request, "sections": db.SECTIONS, "selected": name, "active": "section"},
    )


# ---------- Public display ----------
@app.get("/display")
def display_view(request: Request):
    return templates.TemplateResponse("display.html", {"request": request, "active": "display"})


# ---------- Feedback ----------
@app.get("/feedback")
def feedback_form(request: Request, client: int = 0):
    return templates.TemplateResponse(
        "feedback.html",
        {"request": request, "sections": db.SECTIONS, "active": "feedback", "submitted": False, "is_client": bool(client)},
    )


@app.post("/feedback")
def feedback_submit(
    request: Request,
    client: int = 0,
    name: str = Form(""),
    section: str = Form(...),
    reason: str = Form(...),
    rating: str = Form(...),
):
    if rating not in ("good", "average", "poor"):
        rating = "average"
    db.create_feedback(section=section, reason=reason, rating=rating, name=name)
    return templates.TemplateResponse(
        "feedback.html",
        {"request": request, "sections": db.SECTIONS, "active": "feedback", "submitted": True, "is_client": bool(client)},
    )


# ---------- Admin login ----------
@app.get("/admin/login")
def admin_login_form(request: Request, error: str = "", next: str = "/admin"):
    if is_logged_in(request):
        return RedirectResponse(url=next)
    return templates.TemplateResponse("login.html", {"request": request, "error": error, "next": next})


@app.post("/admin/login")
def admin_login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    next: str = Form("/admin"),
):
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        request.session["is_admin"] = True
        return RedirectResponse(url=next, status_code=303)
    return RedirectResponse(url=f"/admin/login?error=1&next={next}", status_code=303)


@app.get("/admin/logout")
def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/")


# ---------- Admin dashboard ----------
@app.get("/admin")
def admin_view(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/admin/login?next=/admin")
    return templates.TemplateResponse(
        "admin.html", {"request": request, "sections": db.SECTIONS, "active": "admin"}
    )


@app.get("/staff")
def staff_redirect():
    return RedirectResponse(url="/admin")


# ---------- Sections management API ----------
@app.get("/api/sections")
def api_get_sections(request: Request):
    if not is_logged_in(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    return JSONResponse(db.get_sections())


@app.post("/api/sections")
async def api_save_sections(request: Request):
    if not is_logged_in(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    sections = body.get("sections", [])
    if not isinstance(sections, list):
        return JSONResponse({"error": "sections must be a list"}, status_code=400)
    seen = set()
    cleaned = []
    for s in sections:
        s = str(s).strip()
        if s and s not in seen:
            seen.add(s)
            cleaned.append(s)
    if not cleaned:
        return JSONResponse({"error": "sections list cannot be empty"}, status_code=400)
    db.save_sections(cleaned)
    return JSONResponse({"ok": True, "sections": cleaned})


# ---------- JSON API ----------
@app.get("/api/clients")
def api_clients(active: bool = False, section: str = None):
    return JSONResponse(db.get_all_clients(active_only=active, section=section))


@app.get("/api/stats")
def api_stats():
    return JSONResponse(db.get_stats())


@app.get("/api/feedback")
def api_feedback(request: Request):
    if not is_logged_in(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    return JSONResponse(db.get_all_feedback())


@app.get("/api/feedback-stats")
def api_feedback_stats(request: Request):
    if not is_logged_in(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    return JSONResponse(db.get_feedback_stats())


@app.post("/api/clients/{client_id}/section/{section}")
def api_update_section(client_id: int, section: str, status: str = Form(...)):
    client = db.update_section_status(client_id, section, status)
    if client is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    return JSONResponse(client)


# ---------- About page ----------
@app.get("/about")
def about_page(request: Request):
    return templates.TemplateResponse("about.html", {"request": request, "active": "about"})


# ---------- Excel download endpoints ----------
@app.get("/api/download/clients")
def download_clients_excel(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/admin/login?next=/admin")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    clients = db.get_all_clients(active_only=False)
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Registration"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F4C47", end_color="0F4C47", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Token", "Name", "Phone", "NIC", "Purpose", "Sections", "Status", "Progress %", "Registered At", "Completed At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, c in enumerate(clients, 2):
        sections_str = ", ".join(c["sections"])
        status_str = ", ".join(f"{s}: {c['section_status'].get(s, 'N/A')}" for s in c["sections"])
        row_data = [
            c["token"], c["name"], c.get("phone", ""), c.get("nic", ""),
            c["purpose"], sections_str, c["overall_status"],
            c["progress"], c["created_at"][:19], c.get("completed_at", "") or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(clients) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"clients_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/download/feedback")
def download_feedback_excel(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/admin/login?next=/admin")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = db.get_all_feedback()
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F7A4F", end_color="1F7A4F", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Date", "Name", "Section", "Reason", "Rating"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    rating_label = {"good": "Good \U0001f60a", "average": "Average \U0001f610", "poor": "Poor \u2639\ufe0f"}
    for i, f in enumerate(items, 2):
        row_data = [
            f["created_at"][:10], f.get("name", "") or "—",
            f["section"], f["reason"], rating_label.get(f["rating"], f["rating"]),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(items) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"feedback_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
